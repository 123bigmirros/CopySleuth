from __future__ import annotations

import io
import hashlib

from fastapi import HTTPException, UploadFile
from PIL import Image
from openpyxl import load_workbook


def load_image_file(file: UploadFile) -> Image.Image:
    try:
        file.file.seek(0)
        hasher = hashlib.sha256()
        while True:
            chunk = file.file.read(1024 * 1024)
            if not chunk:
                break
            hasher.update(chunk)
        digest = hasher.hexdigest()
        file.file.seek(0)
        image = Image.open(file.file).convert("RGB")
        image.info["cache_key"] = f"sha256:{digest}"
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid image: {file.filename}") from exc
    finally:
        file.file.seek(0)
    return image


def load_image_bytes(payload: bytes, filename: str) -> Image.Image:
    try:
        digest = hashlib.sha256(payload).hexdigest()
        image = Image.open(io.BytesIO(payload)).convert("RGB")
        image.info["cache_key"] = f"sha256:{digest}"
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid image: {filename}") from exc
    return image


def load_excel_queries(
    payload: bytes, filename: str
) -> list[tuple[int, str | None, Image.Image]]:
    try:
        workbook = load_workbook(io.BytesIO(payload), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid excel: {filename}") from exc
    worksheet = workbook.active
    texts_by_row: dict[int, str] = {}
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2
    ):
        value = row[0] if row else None
        if value is None:
            continue
        text = str(value).strip()
        if text:
            texts_by_row[row_index] = text

    images_by_row: dict[int, list[Image.Image]] = {}
    for image in getattr(worksheet, "_images", []):
        anchor = getattr(image, "anchor", None)
        start = getattr(anchor, "_from", None)
        if start is None:
            continue
        col_index = getattr(start, "col", None)
        if col_index is None or col_index != 1:
            continue
        row_index = getattr(start, "row", None)
        if row_index is None or row_index < 0:
            continue
        data_loader = getattr(image, "_data", None)
        if data_loader is None:
            continue
        try:
            image_bytes = data_loader()
            pil_image = Image.open(io.BytesIO(image_bytes)).convert("RGB")
            digest = hashlib.sha256(image_bytes).hexdigest()
            pil_image.info["cache_key"] = f"sha256:{digest}"
        except Exception:
            continue
        images_by_row.setdefault(row_index + 1, []).append(pil_image)

    rows: list[tuple[int, str | None, Image.Image]] = []
    for row_index in sorted(images_by_row):
        if row_index <= 1:
            continue
        text = texts_by_row.get(row_index)
        for image in images_by_row[row_index]:
            rows.append((row_index, text, image))
    if not rows:
        raise HTTPException(
            status_code=400,
            detail="Excel contains no images in column B",
        )
    return rows


def load_excel_images(payload: bytes, filename: str) -> list[tuple[int, Image.Image]]:
    queries = load_excel_queries(payload, filename)
    return [(row_index, image) for row_index, _text, image in queries]


def load_excel_keywords(payload: bytes, filename: str) -> list[str]:
    try:
        workbook = load_workbook(io.BytesIO(payload), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid excel: {filename}") from exc
    worksheet = workbook.active
    keywords: list[str] = []
    seen: set[str] = set()
    for row in worksheet.iter_rows(min_row=2, max_col=1, values_only=True):
        value = row[0] if row else None
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        if text in seen:
            continue
        seen.add(text)
        keywords.append(text)
    return keywords
