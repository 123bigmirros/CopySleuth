#!/usr/bin/env python3
import argparse
import json
import os
import sys

import pandas as pd

from video_ocr_pipeline import run_video_ocr


def _normalize_cell(value):
    if pd.isna(value):
        return None
    text = str(value).strip()
    if not text:
        return None
    return text


def _resolve_openpyxl_col(ws, keyword_col, drop_header):
    import openpyxl

    if keyword_col is None:
        # Pick first non-empty column.
        for col_idx in range(1, ws.max_column + 1):
            has_value = False
            for row in range(1, ws.max_row + 1):
                value = ws.cell(row=row, column=col_idx).value
                if value is not None and str(value).strip():
                    has_value = True
                    break
            if has_value:
                return col_idx, 1
        return None, None
    if isinstance(keyword_col, int):
        return keyword_col + 1, 1
    # Try matching header name in first row if requested.
    if drop_header:
        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_idx).value
            if header is not None and str(header).strip() == str(keyword_col):
                return col_idx, 2
    # Fallback: treat as Excel column letter.
    col_letter = str(keyword_col).strip().upper()
    if col_letter.isalpha():
        return openpyxl.utils.column_index_from_string(col_letter), 1
    return None, None


def _load_keywords_with_openpyxl(path, drop_header, keyword_col):
    try:
        import openpyxl
    except Exception as exc:
        raise RuntimeError("openpyxl is required to read Excel files.") from exc

    wb = openpyxl.load_workbook(path, data_only=True)
    keywords = []
    for ws in wb.worksheets:
        col_idx, start_row = _resolve_openpyxl_col(ws, keyword_col, drop_header)
        if col_idx is None:
            continue
        for row in range(start_row, ws.max_row + 1):
            value = ws.cell(row=row, column=col_idx).value
            if value is None:
                continue
            text = str(value).strip()
            if text:
                keywords.append(text)
    return keywords


def _load_keywords_from_excel(path, drop_header, keyword_col):
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel not found: {path}")

    header = 0 if drop_header else None
    sheets = pd.read_excel(path, sheet_name=None, header=header, engine="openpyxl")

    keywords = []
    for _, df in sheets.items():
        if df is None or df.shape[1] == 0:
            continue
        if keyword_col is not None:
            if isinstance(keyword_col, int):
                if keyword_col < 0 or keyword_col >= df.shape[1]:
                    continue
                series = df.iloc[:, keyword_col]
            else:
                if keyword_col not in df.columns:
                    continue
                series = df[keyword_col]
        else:
            series = None
            for idx in range(df.shape[1]):
                candidate = df.iloc[:, idx]
                if any(_normalize_cell(v) for v in candidate):
                    series = candidate
                    break
            if series is None:
                continue
        for value in series:
            text = _normalize_cell(value)
            if text:
                keywords.append(text)

    if not keywords:
        # Pandas can return empty values for formula cells; fallback to openpyxl data_only.
        keywords = _load_keywords_with_openpyxl(path, drop_header, keyword_col)

    seen = set()
    ordered = []
    for kw in keywords:
        if kw not in seen:
            seen.add(kw)
            ordered.append(kw)
    return ordered


def main():
    parser = argparse.ArgumentParser(
        description="Load keywords from Excel (first column of each sheet) and run video OCR."
    )
    parser.add_argument("video", help="Path to input video")
    parser.add_argument("excel", help="Path to Excel file")
    parser.add_argument(
        "--drop-header",
        action="store_true",
        help="Treat first row as header and drop it from keyword extraction",
    )
    parser.add_argument(
        "--keyword-col",
        default=None,
        help="Keyword column: 0-based index or column name (default: first non-empty column)",
    )
    parser.add_argument(
        "--det-model-dir",
        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "PP-OCRv5_server_det_infer"),
        help="Path to detection model directory",
    )
    parser.add_argument(
        "--rec-model-dir",
        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "PP-OCRv5_server_rec_infer"),
        help="Path to recognition model directory",
    )
    parser.add_argument(
        "--textline-model-dir",
        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "PP-LCNet_x0_25_textline_ori_infer"),
        help="Path to textline orientation model directory (if enabled)",
    )
    parser.add_argument("--device", default="gpu", choices=["gpu", "cpu"], help="Inference device")
    parser.add_argument("--lang", default=None, help="Language hint (e.g., 'en', 'ch')")
    parser.add_argument("--use-textline-orientation", action="store_true", help="Enable text line orientation")
    parser.add_argument(
        "--batch-size",
        type=int,
        default=8,
        help="Frames per OCR batch (default: 8)",
    )
    parser.add_argument(
        "--rec-batch-size",
        type=int,
        default=None,
        help="Text recognition batch size (optional, e.g., 16 or 32)",
    )
    parser.add_argument(
        "--frame-stride",
        type=int,
        default=1,
        help="Process every Nth frame (default: 1 = all frames)",
    )
    parser.add_argument(
        "--max-dim",
        type=int,
        default=None,
        help="Resize frames so the longest side is <= this value before OCR",
    )
    parser.add_argument(
        "--frame-select",
        default="stride",
        choices=["stride", "adaptive", "scene", "scenedetect"],
        help="Frame selection mode: stride, adaptive, scene, or scenedetect (default: stride)",
    )
    parser.add_argument(
        "--scene-threshold",
        type=float,
        default=8.0,
        help="Scene/adaptive selection threshold; higher keeps fewer frames (default: 8.0)",
    )
    parser.add_argument(
        "--scenedetect-threshold",
        type=float,
        default=27.0,
        help="PySceneDetect content threshold; lower = more cuts (default: 27.0)",
    )
    parser.add_argument(
        "--dedup-hash-threshold",
        type=int,
        default=8,
        help="dHash distance threshold for near-duplicate removal in scenedetect mode (default: 8)",
    )
    parser.add_argument(
        "--dedup-max-skip-frames",
        type=int,
        default=0,
        help="Force keep at least every N frames in scenedetect dedup; 0 disables (default: 0)",
    )
    parser.add_argument(
        "--max-skip-frames",
        type=int,
        default=60,
        help="Force keep at least every N frames in adaptive/scene mode (default: 60)",
    )
    parser.add_argument(
        "--select-max-dim",
        type=int,
        default=160,
        help="Downscale frames to this max dim for selection (default: 160)",
    )
    parser.add_argument(
        "--no-progress",
        action="store_true",
        help="Disable progress bar",
    )
    parser.add_argument(
        "--max-gap-frames",
        type=int,
        default=3,
        help="Maximum frame gap to merge matches into one segment (default: 3)",
    )
    parser.add_argument("--out", default=None, help="Optional path to save JSON output")
    args = parser.parse_args()

    keyword_col = None
    if args.keyword_col is not None:
        try:
            keyword_col = int(args.keyword_col)
        except ValueError:
            keyword_col = args.keyword_col

    try:
        keywords = _load_keywords_from_excel(args.excel, args.drop_header, keyword_col)
    except (FileNotFoundError, ValueError) as exc:
        print(str(exc), file=sys.stderr)
        return 2

    if not keywords:
        print("No keywords found in Excel.", file=sys.stderr)
        return 2

    try:
        output = run_video_ocr(
            args.video,
            keywords,
            det_model_dir=args.det_model_dir,
            rec_model_dir=args.rec_model_dir,
            textline_model_dir=args.textline_model_dir,
            device=args.device,
            lang=args.lang,
            use_textline_orientation=args.use_textline_orientation,
            max_gap_frames=args.max_gap_frames,
            show_progress=not args.no_progress,
            batch_size=args.batch_size,
            rec_batch_size=args.rec_batch_size,
            frame_stride=args.frame_stride,
            max_dim=args.max_dim,
            frame_select=args.frame_select,
            scene_threshold=args.scene_threshold,
            scenedetect_threshold=args.scenedetect_threshold,
            dedup_hash_threshold=args.dedup_hash_threshold,
            dedup_max_skip_frames=args.dedup_max_skip_frames,
            max_skip_frames=args.max_skip_frames,
            select_max_dim=args.select_max_dim,
        )
    except (FileNotFoundError, ValueError, RuntimeError) as exc:
        print(str(exc), file=sys.stderr)
        return 2

    output["excel"] = args.excel
    output["keyword_count"] = len(keywords)

    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=True, indent=2)
            f.write("\n")
    else:
        json.dump(output, sys.stdout, ensure_ascii=True, indent=2)
        sys.stdout.write("\n")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
